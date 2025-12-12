"""
╔════════════════════════════════════════════════════════════════════════╗
║                    入荷検査フォーム システム                             ║
║                                                                        ║
║  バージョン: v3.3                                                       ║
║  【v3.3 修正内容】                                                     ║
║  ✅ 元のマニュアルフォーマットに直接書き込み                            ║
║  ✅ 「□可　□否」→「☑可」「☑否」に書き換え                            ║
║  ✅ 写真は別シートにカテゴリ別で配置                                    ║
║                                                                        ║
╚════════════════════════════════════════════════════════════════════════╝
"""

import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from datetime import datetime
import json
import os
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
from email.header import Header
from pathlib import Path
from PIL import Image as PILImage
from io import BytesIO
import unicodedata
import copy
import re

# ========== 【 設定・定数 】==========
MANUAL_FILE = "manual.xlsx"
MASTER_FILE = "inspector_master.xlsx"
PHOTO_DIR = "photos"
CONFIG_FILE = "app_config.json"

Path(PHOTO_DIR).mkdir(parents=True, exist_ok=True)

# ========== 【 セッション状態の初期化 】==========
if 'inspection_data' not in st.session_state:
    st.session_state.inspection_data = {}
if 'selected_emails' not in st.session_state:
    st.session_state.selected_emails = []
if 'uploaded_photos' not in st.session_state:
    st.session_state.uploaded_photos = {}
if 'photo_bytes' not in st.session_state:
    st.session_state.photo_bytes = {}
if 'excel_data' not in st.session_state:
    st.session_state.excel_data = None

# ========== 【 ユーティリティ関数 】==========

def normalize_text(text):
    """全角英数字・記号を半角に変換"""
    if text is None:
        return ""
    return unicodedata.normalize('NFKC', str(text))

def normalize_email(email):
    """メールアドレスの全角文字を半角に変換"""
    if email is None:
        return ""
    normalized = unicodedata.normalize('NFKC', str(email))
    normalized = normalized.strip().replace(" ", "").replace("　", "")
    return normalized

# ========== 【 関数定義 】==========

def load_manual():
    """入荷検査マニュアル Excel を読み込み、検査項目を抽出"""
    try:
        wb = openpyxl.load_workbook(MANUAL_FILE)
        ws = wb.worksheets[0]
        
        items = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=11, max_row=45, values_only=False), 1):
            
            if row_idx in [30, 31]:
                continue

            category_cell = row[0]
            description_cell = row[3]
            
            row_content = ""
            for cell in row:
                if cell.value is not None:
                    row_content += str(cell.value).strip() 

            EXCLUDE_KEYWORDS = ["作製部署", "作成部署", "作成者", "作製者", "制定日", "改訂日", "版数", "承認"]
            
            cleaned_row_content = (
                row_content
                .replace(" ", "")
                .replace("　", "")
                .replace("：", "")
                .replace(":", "")
            )

            is_excluded = False
            for keyword in EXCLUDE_KEYWORDS:
                if keyword in cleaned_row_content:
                    is_excluded = True
                    break

            if is_excluded:
                continue
            
            if category_cell.value or description_cell.value:
                category = category_cell.value or ""
                description = description_cell.value or ""
                
                if str(description).strip():
                    # 実際のExcel行番号を保存（min_row=11なので、row_idx + 10）
                    actual_row = row_idx + 10
                    items.append({
                        'id': f"item_{row_idx}",
                        'category': str(category).strip(),
                        'description': str(description).strip(),
                        'row': row_idx,
                        'excel_row': actual_row
                    })
        
        return items

    except Exception as e:
        st.error(f"マニュアル読込エラー: {e}")
        return []
        
def load_masters():
    """検査者マスター Excel を読み込み"""
    try:
        df = pd.read_excel(MASTER_FILE, sheet_name="検査者一覧")
        if 'メールアドレス' in df.columns:
            df['メールアドレス'] = df['メールアドレス'].apply(normalize_email)
        return df
    except Exception as e:
        st.error(f"❌ マスター読込エラー: {e}")
        return pd.DataFrame()

def save_config(emails):
    """メール送信先を保存"""
    try:
        with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
            json.dump({'selected_emails': emails}, f, ensure_ascii=False)
    except Exception as e:
        st.warning(f"⚠️ 設定保存エラー: {e}")

def create_excel_report(inspection_data, photo_bytes, manual_items, writer_name, reviewer_name, inspector_id, lot_no, in_no, inspection_date):
    """
    元のマニュアルフォーマットに検査結果を書き込み
    写真は別シートに配置
    """
    try:
        # 元のマニュアルを読み込み
        wb = openpyxl.load_workbook(MANUAL_FILE)
        ws = wb.worksheets[0]
        
        # ========== ヘッダー情報を書き込み ==========
        # IN.no (B4セル付近を探す)
        ws['B4'] = in_no
        # OR.no (O4セル付近)
        ws['O4'] = ""  # OR.noがあれば
        # 本体S/N (B5セル付近)
        ws['B5'] = inspector_id
        # ロットNo (O5セル付近)
        ws['O5'] = lot_no
        # 入荷日 (B6セル付近)
        ws['B6'] = str(inspection_date)
        # 検査日 (O6セル付近)
        ws['O6'] = str(inspection_date)
        
        # 作業者印・確認者印 (U1, W1付近)
        # 位置は元のフォーマットに合わせて調整が必要
        
        # ========== 検査結果を書き込み ==========
        # V, W, X, Y列に「□可　　　□否」があるので、結果に応じて書き換え
        
        for item in manual_items:
            item_id = item['id']
            excel_row = item['excel_row']
            
            if item_id in inspection_data:
                is_pass = inspection_data[item_id].get('pass', True)
                
                # V列（22列目）の内容を確認して書き換え
                # 元のセルを探す（V列 = 22）
                for col in range(21, 26):  # U, V, W, X, Y列をチェック
                    cell = ws.cell(row=excel_row, column=col)
                    if cell.value:
                        cell_value = str(cell.value)
                        if '□可' in cell_value or '□否' in cell_value:
                            if is_pass:
                                # □可 → ☑可、□否 → □否
                                new_value = cell_value.replace('□可', '☑可')
                            else:
                                # □可 → □可、□否 → ☑否
                                new_value = cell_value.replace('□否', '☑否')
                            cell.value = new_value
                            break
        
        # ========== 写真シートを作成 ==========
        if photo_bytes:
            # 新しいシートを作成
            ws_photo = wb.create_sheet(title="検査写真")
            
            # ヘッダー
            ws_photo['A1'] = "検査写真一覧"
            ws_photo['A1'].font = Font(bold=True, size=16)
            ws_photo.merge_cells('A1:D1')
            
            ws_photo['A3'] = "No."
            ws_photo['B3'] = "カテゴリ"
            ws_photo['C3'] = "検査項目"
            ws_photo['D3'] = "写真"
            
            for cell in ['A3', 'B3', 'C3', 'D3']:
                ws_photo[cell].font = Font(bold=True)
                ws_photo[cell].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                ws_photo[cell].font = Font(bold=True, color="FFFFFF")
            
            # 列幅設定
            ws_photo.column_dimensions['A'].width = 6
            ws_photo.column_dimensions['B'].width = 15
            ws_photo.column_dimensions['C'].width = 40
            ws_photo.column_dimensions['D'].width = 30
            
            row = 4
            photo_count = 0
            
            for idx, item in enumerate(manual_items):
                item_id = item['id']
                
                if item_id in photo_bytes and photo_bytes[item_id]:
                    photo_count += 1
                    
                    ws_photo[f'A{row}'] = photo_count
                    ws_photo[f'B{row}'] = item['category']
                    ws_photo[f'C{row}'] = item['description'][:50]
                    
                    try:
                        # 画像を処理
                        img_data = BytesIO(photo_bytes[item_id])
                        img = PILImage.open(img_data)
                        
                        # 画像をリサイズ（幅150pxに）
                        max_width = 150
                        ratio = max_width / img.width
                        new_height = int(img.height * ratio)
                        img = img.resize((max_width, new_height))
                        
                        # BytesIOに保存
                        img_buffer = BytesIO()
                        img.save(img_buffer, format='PNG')
                        img_buffer.seek(0)
                        
                        # Excelに埋め込み
                        xl_img = XLImage(img_buffer)
                        ws_photo.add_image(xl_img, f'D{row}')
                        
                        # 行の高さを調整
                        ws_photo.row_dimensions[row].height = max(new_height * 0.75, 100)
                        
                    except Exception as img_error:
                        ws_photo[f'D{row}'] = f"写真読込エラー: {img_error}"
                    
                    row += 1
            
            if photo_count == 0:
                ws_photo['A4'] = "写真はありません"
        
        # メモリに保存
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
        
    except Exception as e:
        st.error(f"❌ Excel 作成エラー: {e}")
        import traceback
        st.error(traceback.format_exc())
        return None

def send_email_smtp(recipient_emails, subject, body, excel_data, filename):
    """SMTP 経由でメール送信（Excel 添付）"""
    try:
        smtp_server = st.secrets.get("SMTP_SERVER")
        smtp_port = st.secrets.get("SMTP_PORT", "587")
        smtp_email = st.secrets.get("SMTP_EMAIL")
        smtp_password = st.secrets.get("SMTP_PASSWORD")
        
        if not all([smtp_server, smtp_email, smtp_password]):
            st.error("""
            ❌ SMTP 設定が見つかりません。
            
            Streamlit Cloud で以下を設定してください：
            - SMTP_SERVER
            - SMTP_PORT
            - SMTP_EMAIL
            - SMTP_PASSWORD
            """)
            return False
        
        smtp_email = normalize_email(smtp_email)
        recipient_emails = [normalize_email(e) for e in recipient_emails]
        
        msg = MIMEMultipart()
        msg['From'] = smtp_email
        msg['To'] = ', '.join(recipient_emails)
        msg['Subject'] = Header(subject, 'utf-8')
        
        msg.attach(MIMEText(body, 'plain', 'utf-8'))
        
        part = MIMEBase('application', 'vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        excel_data.seek(0)
        part.set_payload(excel_data.read())
        encoders.encode_base64(part)
        
        part.add_header(
            'Content-Disposition',
            'attachment',
            filename=filename
        )
        msg.attach(part)
        
        with smtplib.SMTP(smtp_server, int(smtp_port)) as server:
            server.starttls()
            server.login(smtp_email, smtp_password)
            server.send_message(msg)
        
        return True
    
    except smtplib.SMTPAuthenticationError:
        st.error("❌ メール認証エラー：パスワード/トークンが間違っています")
        return False
    except smtplib.SMTPException as e:
        st.error(f"❌ SMTP エラー: {e}")
        return False
    except Exception as e:
        st.error(f"❌ メール送信エラー: {type(e).__name__}: {e}")
        return False

# ========== 【 UI・ページレイアウト 】==========

st.set_page_config(page_title="入荷検査フォーム", layout="wide")
st.title("🔍 入荷検査フォーム")

# ========== 【 サイドバー 】==========
with st.sidebar:
    st.header("⚙️ 設定")
    
    masters = load_masters()
    if not masters.empty:
        writer_names = masters['氏名'].tolist()
        emails_list = masters['メールアドレス'].tolist()
        
        st.subheader("👤 作業者情報")
        writer_name = st.selectbox("作業者名", writer_names, key="writer")
        reviewer_name = st.selectbox("確認者名", writer_names, key="reviewer")
        
        st.subheader("📧 メール送信先")
        st.caption("（Excel 確認後に送信する場合のみ選択）")
        selected_emails = st.multiselect(
            "送信先メールアドレス",
            emails_list,
            key="selected_emails"
        )
        
        if selected_emails:
            save_config(selected_emails)
    else:
        st.error("❌ 検査者マスターが見つかりません")
        writer_name = reviewer_name = None
        selected_emails = []
    
    st.subheader("📋 検査情報")
    inspector_id = st.text_input("本体S/N", placeholder="例: SN12345")
    in_no = st.text_input("IN.NO", placeholder="例: IN001")
    lot_no = st.text_input("ロットNO", placeholder="例: LOT001")
    inspection_date = st.date_input("検査日", value=datetime.now())

# ========== 【 メインコンテンツ 】==========
manual_items = load_manual()

if not manual_items:
    st.error("❌ 検査マニュアルの読み込みに失敗しました")
else:
    st.info(f"✅ {len(manual_items)}件の検査項目を読み込みました")
    
    tabs = st.tabs(["検査入力", "確認・送信"])
    
    # ========== 【 TAB 1：検査入力 】==========
    with tabs[0]:
        st.subheader("検査項目入力")
        st.caption("各項目について「可」または「否」を選択してください")
        
        for idx, item in enumerate(manual_items):
            with st.container():
                st.markdown(f"### No. {idx+1}: {item['category']}")
                st.write(f"📝 {item['description']}")
                
                col_check, col_photo = st.columns([2, 3])
                
                with col_check:
                    result = st.radio(
                        f"判定_{item['id']}",
                        ["可", "否"],
                        horizontal=True,
                        label_visibility="collapsed",
                        key=f"result_{item['id']}"
                    )
                    st.session_state.inspection_data[item['id']] = {
                        'description': item['description'],
                        'pass': result == "可",
                        'category': item['category']
                    }
                
                with col_photo:
                    photo = st.file_uploader(
                        f"写真アップロード_{item['id']}",
                        type=['jpg', 'jpeg', 'png'],
                        label_visibility="collapsed",
                        key=f"photo_{item['id']}"
                    )
                    
                    if photo:
                        photo_data = photo.getvalue()
                        st.session_state.photo_bytes[item['id']] = photo_data
                        st.session_state.uploaded_photos[item['id']] = photo.name
                        st.success(f"✅ 写真保存：{photo.name}")
                        
                        img = PILImage.open(BytesIO(photo_data))
                        st.image(img, width=200)
                
                st.divider()
    
    # ========== 【 TAB 2：確認・送信 】==========
    with tabs[1]:
        st.subheader("検査結果確認・送信")
        st.caption("①Excel を確認 → ②メール送信 の流れで進めてください")
        
        if st.session_state.inspection_data:
            col1, col2, col3, col4 = st.columns(4)
            
            with col1:
                passed = sum(1 for v in st.session_state.inspection_data.values() if v.get('pass'))
                st.metric("合格項目", passed)
            
            with col2:
                failed = len(st.session_state.inspection_data) - passed
                st.metric("不合格項目", failed)
            
            with col3:
                photos = len(st.session_state.photo_bytes)
                st.metric("写真添付数", photos)
            
            with col4:
                st.metric("本体S/N", inspector_id if inspector_id else "-")
            
            st.divider()
            
            st.subheader("📊 検査結果一覧")
            result_df = []
            for idx, (item_id, data) in enumerate(st.session_state.inspection_data.items(), 1):
                result_df.append({
                    'No.': idx,
                    'カテゴリ': data['category'],
                    '検査項目': data['description'][:50],
                    '判定': "✅ 可" if data['pass'] else "❌ 否",
                    '写真': "📷 あり" if item_id in st.session_state.photo_bytes else "なし"
                })
            
            result_table = pd.DataFrame(result_df)
            st.dataframe(result_table, use_container_width=True)
            
            st.divider()
            
            # ========== 【 ステップ 1：Excel 生成・ダウンロード 】==========
            st.subheader("💾 ステップ 1️⃣：Excel 生成・確認")
            st.caption("元のマニュアルフォーマットに結果を書き込み、写真は別シートに配置します")
            
            if st.button("📊 Excel を生成・ダウンロード", use_container_width=True):
                if writer_name and reviewer_name:
                    excel_data = create_excel_report(
                        st.session_state.inspection_data,
                        st.session_state.photo_bytes,
                        manual_items,
                        writer_name, reviewer_name, inspector_id,
                        lot_no, in_no, inspection_date
                    )
                    if excel_data:
                        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                        filename = f"inspection_{timestamp}.xlsx"
                        st.session_state.excel_data = excel_data
                        st.session_state.excel_filename = filename
                        
                        st.download_button(
                            label="📥 Excel をダウンロード",
                            data=excel_data,
                            file_name=filename,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                        st.success(f"✅ Excel 生成完了：{filename}")
                        st.info("📋 シート1: 検査結果（元のフォーマット）\n📷 シート2: 検査写真")
                else:
                    st.error("❌ 作業者名と確認者名を選択してください")
            
            st.divider()
            
            # ========== 【 ステップ 2：メール送信 】==========
            st.subheader("📧 ステップ 2️⃣：メール送信")
            
            if selected_emails and st.session_state.excel_data:
                st.info(f"📬 送信先： {len(selected_emails)}件 選択済み")
                
                if st.button("📮 検査結果をメール送信", use_container_width=True, key="send_email_btn"):
                    with st.spinner("📧 メール送信中..."):
                        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                        filename = f"inspection_{timestamp}.xlsx"
                        
                        subject = f"Inspection Result - {in_no} / {lot_no}"
                        body = f"""
入荷検査が完了しました。

【検査情報】
本体S/N: {inspector_id}
IN.NO: {in_no}
ロットNO: {lot_no}
作業者: {writer_name}
確認者: {reviewer_name}
検査日: {inspection_date}

【結果】
合格項目: {passed}件
不合格項目: {failed}件

詳細は添付の Excel ファイルをご確認ください。
- シート1: 検査結果（元のフォーマット）
- シート2: 検査写真

---
入荷検査フォーム v3.3
"""
                        
                        st.session_state.excel_data.seek(0)
                        
                        success = send_email_smtp(
                            selected_emails,
                            subject,
                            body,
                            st.session_state.excel_data,
                            filename
                        )
                        
                        if success:
                            st.success(f"✅ メール送信完了！")
            
            elif not selected_emails:
                st.info("📧 メール送信をご希望の場合は、サイドバーで送信先を選択してください")
            elif not st.session_state.excel_data:
                st.info("📊 先に「Excel を生成・ダウンロード」を実行してください")
        
        else:
            st.info("ℹ️ 検査項目に回答してから「確認・送信」タブをご覧ください")

st.divider()
st.caption("入荷検査フォーム v3.3 | 元フォーマット対応版")
